"""
Procesador de Subitems

Herramienta con interfaz gráfica para limpiar y organizar
exportaciones de Monday.com en formato Excel. Las filas de subitems
se completan automáticamente, se eliminan encabezados repetidos y se
resaltan los cambios en el archivo resultante.
"""

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, Toplevel, Label
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageTk
import os
import sys

BASE_PATH = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))

def procesar_archivo():
    """Procesa el archivo seleccionado y genera una copia limpiada.

    Pasos principales:
    1. Solicitar el archivo al usuario mediante un cuadro de diálogo.
    2. Completar filas de subitems utilizando la última fila válida.
    3. Remover encabezados y filas que no corresponden al reporte final.
    4. Guardar un nuevo Excel con colores indicativos y fechas formateadas.
    """

    archivo = filedialog.askopenfilename(
        title="Selecciona tu archivo Excel",
        filetypes=[("Archivos Excel", "*.xlsx")]
    )
    if not archivo:
        return

    cargando = Toplevel(ventana)
    cargando.title("Procesando archivo...")
    cargando.geometry("300x100")
    cargando.configure(bg="#1e1e1e")
    # Pequeña ventana para informar que la tarea está en progreso
    Label(cargando, text="⏳ Procesando, por favor espera...", bg="#1e1e1e", fg="white", font=("Segoe UI", 11)).pack(expand=True)
    cargando.update()

    try:
        # Cargar el Excel ignorando las dos primeras filas de encabezado
        df = pd.read_excel(archivo, header=2)
        # Columnas auxiliares para identificar cada fila y su estado
        df['__original_index__'] = df.index
        df['__color__'] = None
        df['__eliminar__'] = False

        # Variables de control utilizadas durante la iteración
        ultima_fila_valida = None
        dentro_de_subitems = False

        # Recorrer todas las filas para completar y marcar información
        for i in range(len(df)):
            if pd.isna(df.iloc[i, 0]):
                valor_columna_a = ''
            else:
                valor_columna_a = str(df.iloc[i, 0])
                valor_columna_a = valor_columna_a.encode('ascii', errors='ignore').decode()
                valor_columna_a = valor_columna_a.strip().lower().replace(" ", "").replace("\xa0", "").replace("\t", "")

            if "subitem" in valor_columna_a:
                # Al detectar el texto "subitem" marcamos la fila para eliminar
                # y recordamos que estamos dentro de un bloque de subitems
                dentro_de_subitems = True
                df.loc[i, '__eliminar__'] = True
                if i > 0:
                    # La fila previa se pintará en azul para indicar que es la original
                    df.loc[i - 1, '__color__'] = "azul"
                continue

            if dentro_de_subitems and pd.isna(df.iloc[i, 0]):
                # Completar la fila vacía copiando datos de la última fila válida
                if ultima_fila_valida is not None:
                    for col in df.columns:
                        if col == 'Quote - SAP':
                            if (pd.isna(df.at[i, col]) or df.at[i, col] == '') and pd.notna(df.iloc[i, 1]):
                                df.at[i, col] = df.iloc[i, 1]
                        elif col == df.columns[2]:
                            df.at[i, col] = ultima_fila_valida[col]
                        elif col == df.columns[4]:
                            df.at[i, col] = ultima_fila_valida[col]
                        elif pd.isna(df.at[i, col]) or df.at[i, col] == '':
                            df.at[i, col] = ultima_fila_valida[col]
                    # Esta fila generada se marcará en amarillo en el Excel final
                    df.loc[i, '__color__'] = "amarillo"
                continue

            dentro_de_subitems = False
            ultima_fila_valida = df.iloc[i].copy()

        # Buscar encabezados duplicados y marcar esos bloques junto con las
        # tres filas anteriores para eliminarlos del resultado final
        encabezado = [
            'Name', 'Subitems', 'RFQ Number', 'Quote - SAP', 'Processed by:', 'Status',
            'Received Date', 'Required Bid Date', 'Submitted Date', 'Factory Input',
            'Accounts', 'Location', 'DO AE', 'Account Name', 'DO #', 'Response Time',
            'Late?', 'ABBGDL Email'
        ]
        filas_a_eliminar = []
        for i in range(len(df)):
            fila = list(df.iloc[i].fillna('').astype(str).str.strip())
            if fila[:len(encabezado)] == encabezado:
                for j in range(i - 3, i + 1):
                    if j >= 0:
                        filas_a_eliminar.append(j)
        df.loc[filas_a_eliminar, '__eliminar__'] = True

        # Quitar la fila de etiquetas de ejemplo que suele aparecer en las
        # exportaciones de Monday.com
        fila_objetivo = ['subitems', 'name', 'owner', 'quote - sap', 'special features']
        for i in range(len(df)):
            fila = list(df.iloc[i].fillna('').astype(str).str.strip().str.lower())
            if fila[:5] == fila_objetivo:
                df.loc[i, '__eliminar__'] = True

        # Remover del DataFrame todas las filas que se marcaron previamente
        df = df[df['__eliminar__'] != True].reset_index(drop=True)

        # Generar el nombre del nuevo archivo sin sobrescribir alguno existente
        base_salida = archivo.replace(".xlsx", "_procesado.xlsx")
        salida = base_salida
        contador = 1
        while os.path.exists(salida):
            salida = base_salida.replace(".xlsx", f" ({contador}).xlsx")
            contador += 1

        df_sin_aux = df.drop(columns=['__original_index__', '__eliminar__', '__color__'])
        df_sin_aux.to_excel(salida, index=False)

        # Abrir el archivo generado para aplicar colores y formatos de fecha
        wb = load_workbook(salida)
        ws = wb.active

        azul = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        amarillo = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

        for i, row in df.iterrows():
            color = row.get('__color__')
            if color == "azul":
                ws[f"A{i + 2}"].fill = azul
            elif color == "amarillo":
                ws[f"A{i + 2}"].fill = amarillo

        columnas_fecha = ['Received Date', 'Required Bid Date', 'Submitted Date']
        col_idx = {}
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value in columnas_fecha:
                col_idx[cell.value] = get_column_letter(idx)

        for col_name, col_letter in col_idx.items():
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                if cell.is_date:
                    cell.number_format = 'yyyy-mm-dd'

        wb.save(salida)
        wb.close()

    finally:
        cargando.destroy()

    os.startfile(salida)
    os.startfile(os.path.dirname(salida))
    messagebox.showinfo("✅ Listo", f"Archivo procesado con éxito:\n{salida}")

# Configuración de la interfaz gráfica principal
ventana = tk.Tk()
ventana.title("Procesador de Subitems")
ventana.geometry("550x350")
ventana.configure(bg="#1e1e1e")

icono_path = os.path.join(BASE_PATH, "logo.ico")
imagen_path = os.path.join(BASE_PATH, "logo.png")

try:
    ventana.iconbitmap(icono_path)
except:
    pass

logo_img = Image.open(imagen_path)
logo_img = logo_img.resize((150, 150))
logo = ImageTk.PhotoImage(logo_img)

label_logo = tk.Label(ventana, image=logo, bg="#1e1e1e")
label_logo.pack(pady=10)

label_titulo = tk.Label(ventana, text="Procesador de Subitems", font=("Segoe UI", 18, "bold"), bg="#1e1e1e", fg="white")
label_titulo.pack(pady=5)

estilo = ttk.Style()
estilo.theme_use("clam")
estilo.configure("TButton", foreground="white", background="#2d2d2d", font=("Segoe UI", 12), padding=10)
estilo.map("TButton", background=[("active", "#3c3c3c")])

boton = ttk.Button(ventana, text="📂 Seleccionar archivo Excel", command=procesar_archivo)
boton.pack(pady=20)

ventana.mainloop()
